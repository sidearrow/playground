import argparse
import csv
import yaml
import os
import openpyxl

from processing_config import ProcessingConfig

XLSX_ERROR_VALUE = [
    '#DIV/0!'
]


def xlsx_error_value_to_none(val):
    return None if val in XLSX_ERROR_VALUE else val


def get_data_from_xlsx(config: ProcessingConfig):
    wb = openpyxl.load_workbook(config.input_file_path)
    sheet = wb[wb.sheetnames[0]]

    res = []
    for row_num in range(config.start_row_index + 1, config.end_row_index + 1):
        row_dict = {}

        is_skip = False
        for i in config.skip_empty_col_indexes:
            if sheet.cell(row=row_num, column=i+1).value == None:
                is_skip = True
                break

        if is_skip:
            continue

        for i, col_name in config.col_mappings.items():
            raw_val = sheet.cell(row=row_num, column=i+1).value
            row_dict[col_name] = xlsx_error_value_to_none(raw_val)

        res.append(row_dict)

    return res


def main(config: ProcessingConfig):
    data = get_data_from_xlsx(config)

    fieldnames = list(config.col_mappings.values())
    with open(config.output_file_path, 'w', encoding='utf-8', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(data)


argparser = argparse.ArgumentParser()
argparser.add_argument('--config-file', dest='config_file', required=True)

args = argparser.parse_args()

config_file_path = os.path.join('./processing_config', args.config_file)
with open(config_file_path, 'r') as f:
    config_dict = yaml.load(f)

processing_config = ProcessingConfig(config_dict)

main(processing_config)
